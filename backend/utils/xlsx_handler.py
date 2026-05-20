"""Excel 导入导出工具"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 表头模糊匹配关键词映射
HEADER_KEYWORDS = {
    "seq": ["序号"],
    "source": ["指标来源", "来源"],
    "assessor": ["评价部门", "主考单位", "评价单位", "考核单位", "牵头单位", "评价部门（科室）", "评价\n部门"],
    "dimension": ["维度", "考核维度", "评价维度", "指标名称", "考核指标"],
    "key_work": ["重点工作", "重点任务", "工作事项", "重点\n工作"],
    "main_task": ["主要任务", "工作任务", "任务名称", "考核内容", "考核事项"],
    "scoring_note": ["评分说明", "评分标准", "考核标准", "计分办法", "计分规则"],
    "unit": ["被考核对象", "被考核单位", "考核对象", "责任单位", "被考核\n对象"],
    "period": ["晾晒周期", "考核周期", "评价周期", "周期", "考核频次", "晾晒\n周期"],
    "remark": ["备注"],
}

# 公文标准字体
FONT_TITLE = "方正小标宋简体"   # 标题：小标宋体
FONT_HEADER = "黑体"            # 表头：黑体
FONT_CONTENT = "仿宋_GB2312"    # 正文：仿宋

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def export_xlsx(headers, rows, merge_col=None):
    """通用导出 xlsx（用于导入模板下载）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    if merge_col is not None and rows:
        _merge_column(ws, merge_col, rows, start_row=2)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_gov_xlsx(title, headers, rows, merge_col=None):
    """公文格式导出：标题(小标宋22pt加粗)+表头(黑体14pt)+正文(仿宋12pt)，居中对齐，细边框"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ncols = len(headers)

    # ---- 标题行（合并居中，小标宋22pt加粗，无边框） ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name=FONT_TITLE, size=22, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    # ---- 表头行（黑体14pt加粗，居中，细边框） ----
    header_font = Font(name=FONT_HEADER, size=14, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ---- 数据行（仿宋12pt，居中，细边框） ----
    content_font = Font(name=FONT_CONTENT, size=12)
    content_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, 3):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = content_font
            cell.alignment = content_align
            cell.border = thin_border

    # ---- 维度列合并 ----
    if merge_col is not None and rows:
        _merge_column(ws, merge_col, rows, start_row=3)

    # ---- 列宽 ----
    col_widths = {
        0: 8,   # 序号
        1: 14,  # 被考核单位/维度
        2: 14,  # 维度/评价部门
        3: 14,  # 评价部门/重点工作
        4: 14,  # 重点工作/主要任务
    }
    # 主要任务列(索引5)给60宽，完成情况/得分/评语等给18
    for col_idx in range(ncols):
        letter = ws.cell(row=2, column=col_idx + 1).column_letter
        w = col_widths.get(col_idx, 60 if col_idx == 5 else 18 if col_idx >= ncols - 3 else 14)
        ws.column_dimensions[letter].width = w

    # 打印设置
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _merge_column(ws, merge_col, rows, start_row):
    """合并指定列中连续相同值的单元格"""
    col = merge_col + 1  # 1-based
    start = start_row
    merge_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in range(1, len(rows)):
        if rows[i][merge_col] != rows[i - 1][merge_col]:
            if start < start_row + i - 1:
                ws.merge_cells(start_row=start, start_column=col, end_row=start_row + i - 1, end_column=col)
                ws.cell(row=start, column=col).alignment = merge_align
            start = start_row + i
    if start < start_row + len(rows) - 1:
        ws.merge_cells(start_row=start, start_column=col, end_row=start_row + len(rows) - 1, end_column=col)
        ws.cell(row=start, column=col).alignment = merge_align


def read_xls_sheets(file_bytes):
    """读取 .xls 旧格式文件，返回 {sheet_name: [row_values], ...}"""
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sheets = {}
    for sname in wb.sheet_names():
        ws = wb.sheet_by_name(sname)
        rows = []
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) if ws.cell_value(r, c) != '' else None
                         for c in range(ws.ncols)])
        sheets[sname] = rows
    return sheets


def read_xlsx_sheets(file_bytes):
    """读取 .xlsx 格式文件，返回 {sheet_name: [row_values], ...}"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheets = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        sheets[sname] = rows
    wb.close()
    return sheets


def read_workbook(file_bytes, filename):
    """通用工作簿读取：自动识别 .xls / .xlsx 格式，返回 {sheet_name: [row_values], ...}"""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext == 'xls':
        try:
            return read_xls_sheets(file_bytes)
        except ImportError:
            return None, "需要安装 xlrd 库来支持 .xls 格式：pip install xlrd"
        except Exception as e:
            return None, f".xls 文件解析失败: {str(e)}"
    else:
        try:
            return read_xlsx_sheets(file_bytes)
        except Exception as e:
            return None, f".xlsx 文件解析失败: {str(e)}"


def fuzzy_match_headers(row_values, target_map):
    """
    模糊表头匹配：根据关键词将实际列索引映射到目标字段。
    target_map: { field_name: [keywords] }
    返回: { field_name: col_index } 或 None（匹配失败时）
    """
    # 清洗表头值
    cleaned = []
    for v in row_values:
        if v is None:
            cleaned.append("")
        else:
            s = str(v).strip().replace('\n', '').replace(' ', '')
            cleaned.append(s)

    mapping = {}
    used_cols = set()
    for field, keywords in target_map.items():
        found = -1
        for col_idx, header_val in enumerate(cleaned):
            if not header_val or col_idx in used_cols:
                continue
            for kw in keywords:
                if kw in header_val:
                    found = col_idx
                    break
            if found >= 0:
                break
        if found >= 0:
            mapping[field] = found
            used_cols.add(found)

    # 至少需要5个核心字段才认为匹配成功
    if len(mapping) < 5:
        return None
    return mapping


def normalize_rows(rows, col_mapping, fill_fields=None):
    """
    根据模糊匹配的列索引，将原始行转换为统一格式。
    col_mapping: { field_name: col_index }
    fill_fields: 需要向前填充的字段列表（用于处理合并单元格）
    返回: [{field_name: value}, ...]
    """
    if fill_fields is None:
        fill_fields = ["dimension"]

    prev = {}
    result = []
    for row in rows:
        item = {}
        for field, col_idx in col_mapping.items():
            val = row[col_idx] if col_idx < len(row) else None
            item[field] = str(val).strip() if val is not None else ""

        # 向前填充指定字段（处理合并单元格）
        for f in fill_fields:
            if not item.get(f) and prev.get(f):
                item[f] = prev[f]

        result.append(item)
        prev = item

    return result


def parse_import(file_stream, expected_headers, fill_cols=None):
    """通用导入 xlsx 解析，fill_cols 指定需要向前填充的列索引列表(0-based)，用于处理合并单元格"""
    try:
        wb = load_workbook(file_stream, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, "文件为空"

        headers = [str(h).strip() if h else "" for h in rows[0]]
        if headers != expected_headers:
            return None, f"表头不匹配，期望: {expected_headers}，实际: {headers}"

        data_rows = []
        prev = [""] * len(expected_headers)
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            values = [str(v).strip() if v is not None else "" for v in row]
            # 向前填充指定列
            if fill_cols:
                for c in fill_cols:
                    if not values[c] and prev[c]:
                        values[c] = prev[c]
            data_rows.append(values)
            prev = values

        return (headers, data_rows)
    except Exception as e:
        return None, f"文件解析失败: {str(e)}"
